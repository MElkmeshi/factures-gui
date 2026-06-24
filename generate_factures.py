#!/usr/bin/env python3
"""
Generate one PDF invoice (facture) per delivery driver from a "Factures Livreurs"
workbook. The visual template (the `Exemple` sheet) is hardcoded into the app
(see factures_template.py), so an uploaded workbook only needs a `Sheet1`. A
workbook that still ships its own `Exemple` sheet keeps using it.

Rules:
  * Only Sheet1 is read for driver data.
  * One invoice per driver who has a non-zero value in at least one
    "Services ..." column. Drivers with all-zero services are skipped.
  * "Numéro de facture" = month index. Counting started in August 2025, so for
    2025: Aug=1, Sep=2, Oct=3, Nov=4, Dec=5. From 2026 on it is the plain
    calendar month (Jan=1, Feb=2, ...). The month is read from the most explicit
    signal available: a "TOTAL <MONTH>" column header, else dd/mm dates in the
    amount headers, else ISO week numbers (headers like "W48").
  * "Date de facture" = latest service date found in the amount headers
    (e.g. "...-25/08") combined with the year; for monthly files that carry no
    per-line date it is the last day of the billing month.

Usage:
  python3 generate_factures.py [workbook.xlsx] [--year 2025] [--out DIR] [--keep-xlsx]

Requires: openpyxl, and LibreOffice (`soffice`) for the PDF conversion.
"""
import argparse
import calendar
import datetime
import re
import shutil
import subprocess
import sys
from copy import copy
from io import BytesIO
from pathlib import Path

import openpyxl
from openpyxl.drawing.image import Image as XLImage

from factures_template import load_template_workbook

MONTHS = {
    "JANUARY": 1, "FEBRUARY": 2, "MARCH": 3, "APRIL": 4, "MAY": 5, "JUNE": 6,
    "JULY": 7, "AUGUST": 8, "SEPTEMBER": 9, "OCTOBER": 10, "NOVEMBER": 11,
    "DECEMBER": 12,
    # French, just in case a future file uses it
    "JANVIER": 1, "FEVRIER": 2, "FÉVRIER": 2, "MARS": 3, "AVRIL": 4, "MAI": 5,
    "JUIN": 6, "JUILLET": 7, "AOUT": 8, "AOÛT": 8, "SEPTEMBRE": 9,
    "OCTOBRE": 10, "NOVEMBRE": 11, "DECEMBRE": 12, "DÉCEMBRE": 12,
}

# Fixed layout coordinates in the `Exemple` template ---------------------------
HEADER_SHEET = "Exemple"
C_INV_NUM = "E4"        # Numéro de facture
C_INV_DATE = "E5"       # Date de facture
C_NAME = "B7"           # Driver name
C_REF = "D7"            # Driver id (Ref)
C_ADDR = "B8"           # Address
C_MF = "B9"             # Matricule fiscal
ITEM_FIRST_ROW = 27     # first line-item row
ITEM_BLANK_ROW = 30     # a blank styled row used to reset unused item rows
R_SUBTOTAL = "E36"
R_TVA = "E39"
R_TIMBRE = "E40"
R_TTC = "E41"
R_FINAL = "E43"
TND_FMT = "#,##0.00\\ [$TND]"  # canonical money format used across the invoice


def find_soffice():
    candidates = (
        "soffice", "libreoffice",
        # macOS
        "/Applications/LibreOffice.app/Contents/MacOS/soffice",
        # Windows
        r"C:\Program Files\LibreOffice\program\soffice.exe",
        r"C:\Program Files (x86)\LibreOffice\program\soffice.exe",
    )
    for cand in candidates:
        path = shutil.which(cand) if "/" not in cand and "\\" not in cand else (cand if Path(cand).exists() else None)
        if path:
            return path
    raise RuntimeError("LibreOffice (soffice) not found. Install it to render PDFs.")


def locate_table(ws):
    """Find the header row and the driver-name column on the data sheet.

    Prefers an exact "Driver Name" header, then falls back to plain "Driver" and
    finally any header containing "driver" (ignoring "Driver ID"), so sheets that
    label the column slightly differently still work.
    """
    predicates = (
        lambda h: h == "driver name",
        lambda h: h == "driver",
        lambda h: "driver" in h and "id" not in h,
        lambda h: "driver" in h,
    )
    for match in predicates:
        for r in range(1, 30):
            for c in range(1, ws.max_column + 1):
                val = ws.cell(r, c).value
                if val and match(str(val).strip().lower()):
                    headers = {c2: (str(ws.cell(r, c2).value).strip() if ws.cell(r, c2).value else "")
                               for c2 in range(1, ws.max_column + 1)}
                    return r, headers, c
    raise ValueError("could not find a 'Driver Name' (or 'Driver') header on the sheet.")


def pick_data_sheet(wb):
    """Return the sheet that holds the driver table.

    Prefers 'Sheet1', then falls back to 'Details' (newer monthly files keep a
    plain driver directory on Sheet1 and the billable table on 'Details'), then
    any other sheet with a recognizable driver header.
    """
    ordered = [n for n in ("Sheet1", "Details") if n in wb.sheetnames]
    ordered += [ws.title for ws in wb.worksheets if ws.title not in ordered]
    for name in ordered:
        try:
            locate_table(wb[name])
            return wb[name]
        except ValueError:
            continue
    raise ValueError("no sheet with a 'Driver'/'Driver Name' header found "
                     "(looked at 'Sheet1', 'Details').")


def col_by(headers, *names):
    for c, h in headers.items():
        if h.lower() in [n.lower() for n in names]:
            return c
    return None


# Header names that identify a driver (not a billable amount); used to locate
# the amount columns on layouts that don't label them "Services ...".
IDENTITY_HEADERS = {
    "driver", "driver name", "driver id", "type", "mf", "cin", "check cin",
    "check", "adresse", "address", "phone", "status", "manual id", "issues",
    "patente/auto entrepreneur", "matriculep", "remarque", "name", "id", "ref",
}


def parse_invoice_meta(headers, year):
    """Derive invoice number and date from the service/total headers."""
    total_hdr = next((h for h in headers.values() if h.upper().startswith("TOTAL")), "")

    # 1) Locate the per-period amount columns.
    service_cols = [c for c, h in headers.items() if h.lower().startswith("services")]
    if not service_cols:
        # Newer layouts ('Details' sheet): the amount columns sit after the
        # identity columns. They may be bounded by a 'Total ...' column (e.g.
        # weeks W44-W47 then 'Total September') or simply run to the end of the
        # table when there is no total column at all (e.g. W48-W51).
        total_col = next((c for c, h in headers.items()
                          if h.upper().startswith("TOTAL")), None)
        last_id = max([c for c, h in headers.items()
                       if h and h.lower() in IDENTITY_HEADERS], default=0)
        upper = total_col if total_col else max(headers) + 1
        service_cols = [c for c, h in headers.items()
                        if h and last_id < c < upper
                        and h.lower() not in IDENTITY_HEADERS]
    if not service_cols:
        raise ValueError("no service/amount columns found on the data sheet.")

    # 2) Determine the billing month, most explicit signal first: a
    #    'Total <MONTH>' name, then dd/mm dates, then ISO week numbers (headers
    #    like 'W48') found in the amount columns.
    dates = [m for m in (re.search(r"(\d{1,2})/(\d{1,2})", headers[c])
                         for c in service_cols) if m]
    weeks = [int(m.group(1)) for c in service_cols
             if (m := re.fullmatch(r"[Ww](\d{1,2})", headers[c].strip()))]
    month = None
    for token in re.split(r"[\s_]+", total_hdr.upper()):
        if token in MONTHS:
            month = MONTHS[token]
            break
    if month is None and dates:
        month = max(int(m.group(2)) for m in dates)
    if month is None and weeks:
        month = datetime.date.fromisocalendar(year, max(weeks), 7).month
    if month is None:
        raise ValueError("could not determine the billing month from the headers "
                         "(no 'Total <month>', dates, or week numbers found).")

    # 3) Invoice number: Aug 2025 = 1 ... Dec 2025 = 5; calendar month from 2026.
    if year == 2025 and month >= 8:
        inv_num = month - 7
    else:
        inv_num = month

    # 4) Invoice date: an explicit dd/mm in the headers wins; otherwise the last
    #    day of the billing month (these monthly files carry no per-line dates).
    best = None
    for m in dates:
        d = datetime.datetime(year, int(m.group(2)), int(m.group(1)))
        best = d if best is None or d > best else best
    if best is None:
        best = datetime.datetime(year, month, calendar.monthrange(year, month)[1])
    inv_date = best
    return service_cols, inv_num, inv_date, total_hdr


def clone_style(src_cell, dst_cell):
    dst_cell.font = copy(src_cell.font)
    dst_cell.fill = copy(src_cell.fill)
    dst_cell.border = copy(src_cell.border)
    dst_cell.alignment = copy(src_cell.alignment)
    dst_cell.number_format = src_cell.number_format


def fill_invoice(ws, drv, services, inv_num, inv_date):
    """Write one driver's data into the template sheet (literal values)."""
    ws[C_INV_NUM] = inv_num
    ws[C_INV_DATE] = inv_date
    ws[C_NAME] = drv["name"]
    ws[C_REF] = drv["id"]
    ws[C_ADDR] = drv["addr"]
    ws[C_MF] = drv["mf"]

    n = len(services)
    # template ships with 3 styled item rows (27-29); style for extra rows is
    # cloned from row 29, unused rows are reset from the blank row 30.
    style_src_row = ITEM_FIRST_ROW + 2
    for i in range(9):  # template has room for up to 9 items before the totals
        r = ITEM_FIRST_ROW + i
        a, b, cc, d, e = (ws.cell(r, col) for col in (1, 2, 3, 4, 5))
        if i < n:
            label, value = services[i]
            if r > style_src_row:
                for col, srcc in zip((1, 2, 3, 4, 5), (1, 2, 3, 4, 5)):
                    clone_style(ws.cell(style_src_row, srcc), ws.cell(r, col))
            a.value, b.value, cc.value, d.value, e.value = i + 1, label, 1, value, value
        else:
            for col, dst in zip((1, 2, 3, 4, 5), (a, b, cc, d, e)):
                clone_style(ws.cell(ITEM_BLANK_ROW, col), dst)
                dst.value = None

    subtotal = sum(v for _, v in services)
    ws[R_SUBTOTAL] = subtotal
    ws[R_TVA] = 0
    ws[R_TIMBRE] = 0
    ws[R_TTC] = subtotal
    ws[R_FINAL] = subtotal

    # the template has a stray euro-formatted spacer cell; force all money
    # cells to the TND format so nothing renders in euros.
    for row in ws.iter_rows():
        for cell in row:
            if cell.number_format and "€" in cell.number_format:
                cell.number_format = TND_FMT


def safe(text):
    return re.sub(r"[^\w\-. ]", "_", str(text)).strip()


def unique_sheet_name(name, drv_id, used):
    """Excel sheet titles: <=31 chars, no : \\ / ? * [ ], unique."""
    base = re.sub(r"[:\\/?*\[\]]", " ", str(name)).strip()[:31] or "Facture"
    title = base
    n = 2
    while title.lower() in {u.lower() for u in used}:
        suffix = f" {n}"
        title = base[:31 - len(suffix)] + suffix
        n += 1
    used.add(title)
    return title


def generate(workbook="Factures Livreurs.xlsx", year=2025, out=None,
             keep_xlsx=False, no_pdf=False, no_excel=False, log=print):
    """Generate the invoices. Returns the output directory (Path).

    `log` is a callback used for progress messages (defaults to print, so the
    CLI behaves as before; the GUI passes its own logger). Errors are raised as
    exceptions so callers can present them however they like.
    """
    src = Path(workbook)
    if not src.exists():
        raise FileNotFoundError(f"file not found: {src}")
    soffice = find_soffice() if not no_pdf else None

    upload = openpyxl.load_workbook(src)

    # Read driver values from a data sheet (Sheet1, else Details). Load with
    # data_only so formula cells — the 'Details' layout computes amounts via
    # VLOOKUP — yield their cached numbers instead of the formula text.
    values = openpyxl.load_workbook(src, data_only=True)
    data = pick_data_sheet(values)

    # The invoice template is hardcoded (see factures_template.py), so the upload
    # only needs a data sheet. A workbook that still ships its own 'Exemple' sheet
    # keeps using it, for backward compatibility.
    if HEADER_SHEET in upload.sheetnames:
        wb = upload
    else:
        wb = load_template_workbook()
    tmpl = wb[HEADER_SHEET]

    # The template logo's image stream is consumed on the first save, so capture
    # its bytes + anchor once and rebuild a fresh image before each save.
    logos = [(img._data(), img.anchor) for img in tmpl._images]

    header_row, headers, c_name = locate_table(data)
    c_id = col_by(headers, "Driver ID")
    c_mf = col_by(headers, "MF")
    c_addr = col_by(headers, "Adresse", "Address")
    service_cols, inv_num, inv_date, total_hdr = parse_invoice_meta(headers, year)

    log(f"Month column: {total_hdr!r} -> facture #{inv_num}, date {inv_date.date()}")
    log(f"Service columns: {[headers[c] for c in service_cols]}")

    # show only the invoice sheet in the rendered PDF
    for ws in wb.worksheets:
        ws.sheet_state = "visible" if ws.title == HEADER_SHEET else "hidden"
    wb.active = wb.sheetnames.index(HEADER_SHEET)

    out_dir = Path(out) if out else src.parent / f"Factures_{safe(total_hdr) or 'OUTPUT'}_{year}"
    out_dir.mkdir(parents=True, exist_ok=True)

    # collect every qualifying driver once
    drivers, skipped = [], 0
    for r in range(header_row + 1, data.max_row + 1):
        name = data.cell(r, c_name).value
        if not name or not str(name).strip():
            continue
        services = []
        for c in service_cols:
            v = data.cell(r, c).value
            services.append((headers[c], float(v) if isinstance(v, (int, float)) else 0.0))
        if sum(v for _, v in services) <= 0:
            skipped += 1
            continue
        drivers.append({
            "name": str(name).strip(),
            "id": data.cell(r, c_id).value if c_id else "",
            "mf": data.cell(r, c_mf).value if c_mf else "",
            "addr": data.cell(r, c_addr).value if c_addr else "",
            "services": [s for s in services if s[1] > 0],  # only non-zero lines
        })

    log(f"Qualifying drivers: {len(drivers)} (skipped {skipped} all-zero drivers).")
    if not drivers:
        log("Nothing to generate.")
        return out_dir

    def refresh_logo(ws):
        ws._images = []
        for raw, anchor in logos:
            img = XLImage(BytesIO(raw))
            img.anchor = anchor
            ws.add_image(img)

    # --- one PDF per driver --------------------------------------------------
    if not no_pdf:
        tmp_dir = out_dir / "_tmp_xlsx"
        tmp_dir.mkdir(exist_ok=True)
        xlsx_files = []
        for drv in drivers:
            fill_invoice(tmpl, drv, drv["services"], inv_num, inv_date)
            refresh_logo(tmpl)
            path = tmp_dir / (safe(f"{inv_num} - {drv['name']} - {drv['id']}") + ".xlsx")
            wb.save(path)
            xlsx_files.append(path)

        log(f"Rendering {len(xlsx_files)} PDFs with LibreOffice…")
        for i in range(0, len(xlsx_files), 25):
            subprocess.run(
                [soffice, "--headless", "--calc", "--convert-to", "pdf",
                 "--outdir", str(out_dir), *map(str, xlsx_files[i:i + 25])],
                check=True, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL,
            )
        log(f"Created {len(list(out_dir.glob('*.pdf')))} PDFs in: {out_dir}")
        if not keep_xlsx:
            shutil.rmtree(tmp_dir)

    # --- one workbook, one sheet per driver ----------------------------------
    if not no_excel:
        book = openpyxl.load_workbook(src) if HEADER_SHEET in upload.sheetnames \
            else load_template_workbook()
        template = book[HEADER_SHEET]
        used = set()
        for drv in drivers:
            sheet = book.copy_worksheet(template)
            sheet.title = unique_sheet_name(drv["name"], drv["id"], used)
            sheet.sheet_state = "visible"
            fill_invoice(sheet, drv, drv["services"], inv_num, inv_date)
            refresh_logo(sheet)
        for name in [s for s in book.sheetnames if s not in
                     {unique for unique in used}]:
            del book[name]
        book.active = 0
        xlsx_out = out_dir / f"Factures_{safe(total_hdr) or 'OUTPUT'}_{year}_par_livreur.xlsx"
        book.save(xlsx_out)
        log(f"Created workbook with {len(drivers)} driver sheets: {xlsx_out}")

    return out_dir


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("workbook", nargs="?", default="Factures Livreurs.xlsx")
    ap.add_argument("--year", type=int, default=2025)
    ap.add_argument("--out", default=None)
    ap.add_argument("--keep-xlsx", action="store_true")
    ap.add_argument("--no-pdf", action="store_true", help="skip the per-driver PDFs")
    ap.add_argument("--no-excel", action="store_true",
                    help="skip the combined workbook with one sheet per driver")
    args = ap.parse_args()
    try:
        generate(args.workbook, year=args.year, out=args.out,
                 keep_xlsx=args.keep_xlsx, no_pdf=args.no_pdf, no_excel=args.no_excel)
    except (FileNotFoundError, ValueError, RuntimeError) as exc:
        sys.exit(f"ERROR: {exc}")


if __name__ == "__main__":
    main()
