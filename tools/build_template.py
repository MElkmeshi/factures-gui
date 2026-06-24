#!/usr/bin/env python3
"""Regenerate factures_template.py from a source workbook's 'Exemple' sheet.

The invoice template (the 'Exemple' sheet, including the Presto logo) is bundled
into the app as a base64 blob so uploaded workbooks only need a 'Sheet1'. Run
this whenever the template design changes:

    python tools/build_template.py "Factures Livreurs.xlsx"

Requires Pillow (so openpyxl preserves the embedded logo on save).
"""
import base64
import io
import sys
import textwrap
from pathlib import Path

import openpyxl

TEMPLATE_SHEET = "Exemple"
OUT = Path(__file__).resolve().parent.parent / "factures_template.py"

MODULE = '''"""Hardcoded invoice template — the 'Exemple' sheet bundled with the app.

Extracted once from 'Factures Livreurs.xlsx'. The driver workbook a user uploads
now only needs a 'Sheet1'; this module supplies the invoice template (layout,
styling and the Presto logo) so it no longer has to travel with the data.

Stored as base64 so it survives the ``*.xlsx`` gitignore (which protects the
sensitive driver data) and ships inside the Docker image with no extra COPY rule.
To regenerate, run ``python tools/build_template.py`` (see README).
"""
import base64
from io import BytesIO

import openpyxl

# The 'Exemple' sheet only — no driver data — as a base64-encoded .xlsx.
_TEMPLATE_XLSX_B64 = """\\
%s
"""


def load_template_workbook():
    """Return a fresh openpyxl workbook holding only the 'Exemple' template sheet."""
    return openpyxl.load_workbook(BytesIO(base64.b64decode(_TEMPLATE_XLSX_B64)))
'''


def main():
    src = Path(sys.argv[1] if len(sys.argv) > 1 else "Factures Livreurs.xlsx")
    if not src.exists():
        sys.exit(f"source workbook not found: {src}")
    try:
        import PIL  # noqa: F401
    except ImportError:
        sys.exit("Pillow is required so the logo survives the openpyxl round-trip "
                 "(pip install Pillow).")

    wb = openpyxl.load_workbook(src)
    if TEMPLATE_SHEET not in wb.sheetnames:
        sys.exit(f"source workbook has no '{TEMPLATE_SHEET}' sheet.")
    for name in [s for s in wb.sheetnames if s != TEMPLATE_SHEET]:
        del wb[name]

    buf = io.BytesIO()
    wb.save(buf)
    b64 = base64.b64encode(buf.getvalue()).decode("ascii")
    OUT.write_text(MODULE % "\n".join(textwrap.wrap(b64, 96)))
    print(f"wrote {OUT} ({len(b64)} base64 chars, logo images: "
          f"{len(wb[TEMPLATE_SHEET]._images)})")


if __name__ == "__main__":
    main()
